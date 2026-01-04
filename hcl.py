import os, json, re
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

URL = "https://www.hclfoundation.org/work-with-us"
HEADERS = {"User-Agent": "Mozilla/5.0"}

# Load keywords
with open("keywords.json", "r") as f:
    keywords = json.load(f)

priority = ["Governance", "Learning", "Safety", "Climate"]

def run_hcl_scraper():
    listings = []

    res = requests.get(URL, headers=HEADERS, timeout=10, verify=False)
    soup = BeautifulSoup(res.text, "html.parser")

    # Find all rows in the opportunities table
    rows = soup.find_all("tr")
    for row in rows:
        title_td = row.find("td", class_="views-field-field-job-title")
        link_td = row.find("td", class_="views-field-field-download-cta")
        deadline_td = row.find("td", class_="views-field-field-post-date")

        if not title_td or not link_td or not deadline_td:
            continue

        title = title_td.get_text(strip=True)

        a_tag = link_td.find("a", href=True)
        link = None
        if a_tag:
            link = a_tag["href"].strip()
            if not link.startswith("http"):
                link = "https://www.hclfoundation.org" + link

        deadline = deadline_td.get_text(strip=True)

        # ✅ Keyword matching (allow multiple verticals)
        text_blob = title.lower()
        matched_verticals = []
        for vertical in priority:
            for word in keywords.get(vertical, []):
                if re.search(r'\b' + re.escape(word.lower()) + r'\b', text_blob):
                    matched_verticals.append(vertical)
                    break  # prevent duplicate entries per vertical

        # ✅ Strict: include only if some keyword matched
        if matched_verticals:
            listings.append({
                "Title": title,
                "Deadline": deadline,
                "Link": link if link else "N/A",
                "Matched_Vertical": ", ".join(matched_verticals)
            })

    if not listings:
        print("⚠️ No matched opportunities found.")
        return

    # Save to Excel
    if not os.path.exists("output"):
        os.makedirs("output")

    df = pd.DataFrame(listings)

    # Parse deadline into date (if possible)
    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, "%d %b %Y").date()
        except:
            try:
                return datetime.strptime(date_str, "%d %B, %Y").date()
            except:
                return date_str  # fallback: keep original string

    df["Deadline_Parsed"] = df["Deadline"].apply(parse_date)

    df["Clickable_Link"] = df.apply(
        lambda row: '=HYPERLINK("{}","{}")'.format(row["Link"], row["Title"].replace('"', '""')),
        axis=1
    )

    # Final dataframe
    df = df[["Title", "Matched_Vertical", "Deadline", "Clickable_Link"]]

    excel_path = "output/hcl_opportunities.xlsx"
    df.to_excel(excel_path, index=False, engine="openpyxl")

    wb = load_workbook(excel_path)
    ws = wb.active
    for col, width in {"A": 80, "B": 30, "C": 20, "D": 80}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(excel_path)

    print(f"✅ HCL Excel saved to {excel_path}")


if __name__ == "__main__":
    run_hcl_scraper()
